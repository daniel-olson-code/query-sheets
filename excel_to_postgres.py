"""
Module for data type conversion and Excel file handling with PostgreSQL integration.

This module provides utilities for converting data types between Python and PostgreSQL,
handling Excel files, and uploading Excel data to PostgreSQL tables.
"""

import datetime
import enum
import itertools
import openpyxl
import psycopg2
import psycopg2.extras
import postgres
import csv
import tqdm
import unsync
import json
import uuid
import contextlib
import tempfile
from typing import Any

import os


@contextlib.contextmanager
def temporary_file_name():
    """
    Context manager for generating a temporary file name.
    The file will be deleted after the context manager exits.
    Yields:
        str: A temporary file name.
    """
    name = f'{uuid.uuid4()}'
    yield name
    try:
        os.remove(name)
    except OSError:
        pass


def jsonl_append(path: str, obj: Any):
    """
    Append an object to a JSON Lines file.
    Args:
        path (str): The path to the JSON Lines file.
        obj (Any): The object to append.
    Returns:
        None
    """
    with open(path, 'a') as f:
        f.write(json.dumps(obj) + '\n')


def jsonl_stream(path: str):
    """
    Context manager for writing to a JSON Lines file.
    Args:
        path (str): The path to the JSON Lines file.
    Yields:
        None
    """
    with open(path) as f:
        for line in f:
            if line.strip():
                yield json.loads(line)


class DataTypes(enum.Enum):
    """Enumeration of PostgreSQL data types."""
    text = 'text'
    real = 'real'
    bigint = 'bigint'
    integer = 'integer'
    json = 'json'
    timestamp = 'timestamp'
    date = 'date'
    boolean = 'boolean'
    # null = 'null'  # <- non-existent. Future plans of adding the feature to choose to cast a curtain data type is "null"


def convert_python_type_to_postgres(val: Any) -> DataTypes:
    """
    Convert a Python value to its corresponding PostgreSQL data type.

    Args:
        val (Any): The Python value to convert.

    Returns:
        DataTypes: The corresponding PostgreSQL data type.
    """
    if val is None:
        return DataTypes.boolean  # DataTypes.null
    if isinstance(val, bool):
        return DataTypes.boolean
    if isinstance(val, str):
        return DataTypes.text
    if isinstance(val, float):
        return DataTypes.real
    if isinstance(val, int):
        if -9223372036854775805 > val or val > 9223372036854775805:
            return DataTypes.text
        if -2147483645 > val or val > 2147483645:
            return DataTypes.bigint
        return DataTypes.integer
    if isinstance(val, (list, dict)):
        return DataTypes.json
    if isinstance(val, datetime.datetime):
        return DataTypes.timestamp
    if isinstance(val, datetime.date):
        return DataTypes.date
    return DataTypes.text


def determine_data_type(previous_type: DataTypes, current_value: Any) -> DataTypes:
    """
    Determine the appropriate data type based on the previous type and current value.

    Args:
        previous_type (DataTypes): The previously determined data type.
        current_value (Any): The current value to check.

    Returns:
        DataTypes: The determined data type.
    """
    current_data_type = convert_python_type_to_postgres(current_value)
    if get_types_casting_hierarchy(current_data_type) < get_types_casting_hierarchy(previous_type):
        return current_data_type
    return previous_type


def get_types_casting_hierarchy(data_type: DataTypes) -> int:
    """
    Get the hierarchy level of a data type for casting purposes.

    Higher values can be cast to lower values, but not necessarily vice versa.

    Args:
        data_type (DataTypes): The data type to get the hierarchy for.

    Returns:
        int: The hierarchy level of the data type (0 is lowest).
    """
    hierarchy = {
        DataTypes.text: 0, DataTypes.real: 1, DataTypes.bigint: 2, DataTypes.integer: 3,
        DataTypes.json: 4, DataTypes.timestamp: 5, DataTypes.date: 6, DataTypes.boolean: 7  # , DataTypes.null: 8
    }
    return hierarchy[data_type] if data_type in hierarchy else 0


def get_sheet_names_xlsx(filepath: str) -> list[str]:
    """
    Get the names of all sheets in an Excel file.

    Args:
        filepath (str): The path to the Excel file.

    Returns:
        list[str]: A list of sheet names in the Excel file.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames


def create_xlsx(table: list[list] | list[dict], path: str, sheet_name: str = 'Sheet') -> None:
    """
    Create an Excel file from a list of lists or list of dictionaries.

    Args:
        table (list[list] | list[dict]): The data to write to the Excel file.
        path (str): The path where the Excel file will be saved.
        sheet_name (str, optional): The name of the sheet in the Excel file. Defaults to 'Sheet'.

    Raises:
        ValueError: If the table is empty.
    """
    if not table:
        raise ValueError('Table cannot be empty')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if isinstance(table[0], dict):
        headers = list(table[0].keys())
        ws.append(headers)

    for row in table:
        if isinstance(row, dict):
            row = list(row.values())
        ws.append(row)

    print('saving to', path)
    wb.save(path)


def xlsx_to_sql(path: str, sheet_name: str, table_name: str, db_params: postgres.DatabaseParameters) -> None:
    """
    Upload data from an Excel file to a PostgreSQL table.

    This function creates a new table in the PostgreSQL database and populates it with data from the Excel file.

    Args:
        path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet in the Excel file to read data from.
        table_name (str): Name of the table to create in the PostgreSQL database.
        db_params (postgres.DatabaseParameters): Connection parameters for the PostgreSQL database.

    Raises:
        psycopg2.Error: If there's an error connecting to the database, creating the table, or executing SQL queries.
    """
    wb = openpyxl.load_workbook(path, read_only=True)  # read_only=True is slower but uses much less memory
    ws = wb[sheet_name]
    ws.reset_dimensions()

    headers = None
    table_data_types = {}

    chunk = []
    chunk_size = 5000

    with tempfile.NamedTemporaryFile() as tmp:
        for row in tqdm.tqdm(ws.iter_rows(values_only=True), desc='Extracting Excel'):
            row = list(row)
            if headers is None:
                headers = row
                continue
            if len(row) < len(headers):
                # Fill missing values with None, extending row length to match header length
                row = itertools.chain(row, itertools.repeat(None, len(headers) - len(row)))
            if len(row) > len(headers):
                # Shorten row to length of headers
                headers = row[:len(headers)]

            # table.append(row)
            # jsonl_append(temp_path, row)
            chunk.append(row)
            if len(chunk) >= chunk_size:
                tmp.write(('\n'.join([json.dumps(r) for r in chunk]) + '\n').encode())
                chunk.clear()

            for h, v in zip(headers, row):
                table_data_types[h] = determine_data_type(table_data_types.get(h, DataTypes.boolean), v)

        if len(chunk) > 0:
            tmp.write(('\n'.join([json.dumps(r) for r in chunk]) + '\n').encode())
            chunk.clear()

        with psycopg2.connect(
                    host=db_params.host,
                    database=db_params.database,
                    user=db_params.user,
                    password=db_params.password,
                    port=db_params.port,
            ) as conn:
            cur = conn.cursor()
            double_quote = '"'
            cur.execute(f'DROP TABLE IF EXISTS "{table_name}"')
            cur.execute(
                f'''CREATE TABLE "{table_name}"
                ({", ".join(f"{double_quote}{h}{double_quote} {v.value}"
                            for h, v in table_data_types.items())})'''
            )

            percent_s = ", ".join(["%s"] * len(headers))
            query = f'INSERT INTO "{table_name}" VALUES ({percent_s})'

            tmp.seek(0)
            for line in tqdm.tqdm(tmp):
                row = json.loads(line.decode())
                chunk.append(tuple(row))
                if len(chunk) >= chunk_size:
                    psycopg2.extras.execute_batch(cur, query, tuple(chunk))
                    chunk.clear()

            if chunk:
                psycopg2.extras.execute_batch(cur, query, tuple(chunk))
                del chunk

            conn.commit()
            cur.close()


def csv_to_sql(
        path: str,
        table_name: str,
        db_params: postgres.DatabaseParameters
) -> None:
    """
    Upload data from a CSV file to a PostgreSQL table.

    This function creates a new table in the PostgreSQL database and populates it with data from the CSV file.

    Args:
        path (str): Path to the CSV file.
        table_name (str): Name of the table to create in the PostgreSQL database.
        db_params (postgres.DatabaseParameters): Connection parameters for the PostgreSQL database.

    Raises:
        psycopg2.Error: If there's an error connecting to the database, creating the table, or executing SQL queries.
    """
    with open(path, 'r') as f:
        reader = csv.reader(f)
        headers = next(reader)
        headers = [h.replace('﻿', '').strip() for h in headers]
        table_data_types = {h: DataTypes.text for h in headers}

        with tempfile.NamedTemporaryFile() as tmp:
            chunk = []
            chunk_size = 5000
            for row in tqdm.tqdm(reader, desc='Extracting CSV'):
                row = list(row)
                if len(row) < len(headers):
                    # Fill missing values with None, extending row length to match header length
                    row = itertools.chain(row, itertools.repeat(None, len(headers) - len(row)))
                if len(row) > len(headers):
                    # Shorten row to length of headers
                    headers = row[:len(headers)]

                # table.append(row)
                # jsonl_append(temp_path, row)
                chunk.append(row)
                if len(chunk) >= chunk_size:
                    tmp.write(('\n'.join([json.dumps(r) for r in chunk]) + '\n').encode())
                    chunk.clear()
                    # tmp.flush()

            if len(chunk) > 0:
                tmp.write(('\n'.join([json.dumps(r) for r in chunk]) + '\n').encode())
                chunk.clear()
                # tmp.flush()

            with psycopg2.connect(
                    host=db_params.host,
                    database=db_params.database,
                    user=db_params.user,
                    password=db_params.password,
                    port=db_params.port,
            ) as conn:
                cur = conn.cursor()
                double_quote = '"'
                cur.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                cur.execute(
                    f'''CREATE TABLE "{table_name}"
                    ({", ".join(f"{double_quote}{h}{double_quote} {v.value}"
                                for h, v in table_data_types.items())})'''
                )
                percent_s = ", ".join(["%s"] * len(headers))
                query = f'INSERT INTO "{table_name}" VALUES ({percent_s})'
                tmp.seek(0)
                for line in tqdm.tqdm(tmp):
                    row = json.loads(line.decode())
                    chunk.append(tuple(row))
                    if len(chunk) >= chunk_size:
                        psycopg2.extras.execute_batch(cur, query, tuple(chunk))
                        chunk.clear()
                        # tmp.flush()
                        # conn.commit()

                if chunk:
                    psycopg2.extras.execute_batch(cur, query, tuple(chunk))
                    del chunk
                    # conn.commit()

                conn.commit()
                cur.close()
